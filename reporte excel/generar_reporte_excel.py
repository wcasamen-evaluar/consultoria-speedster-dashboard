"""Genera el reporte general de evaluaciones en formato Excel.

El archivo consolida Desempeño, Competencias y Objetivos con las mismas
reglas de cálculo y emparejamiento utilizadas por el dashboard. Cuando una
persona no ha realizado una evaluación, la celda correspondiente queda vacía.
"""

from __future__ import annotations

import argparse
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


CARPETA_SCRIPT = Path(__file__).resolve().parent
RAIZ_PROYECTO = CARPETA_SCRIPT.parent
SALIDA_PREDETERMINADA = CARPETA_SCRIPT / "reporte general_excel.xlsx"

# Permite ejecutar el archivo directamente desde cualquier carpeta.
if str(RAIZ_PROYECTO) not in sys.path:
    sys.path.insert(0, str(RAIZ_PROYECTO))

from reporte import calculos as motor_360  # noqa: E402
from reporte import integrado as motor_integrado  # noqa: E402
from reporte import objetivos as motor_objetivos  # noqa: E402
from reporte import potencial as motor_competencias  # noqa: E402


COLUMNAS_SALIDA = ["Nombre", "correo", "desempeño", "competencias", "objetivos"]
COLUMNAS_PUNTAJE = ["desempeño", "competencias", "objetivos"]
COLUMNAS_CORREO = [
    "email_colaborador_360",
    "email_colaborador_obj",
    "correo",
    "correo_potencial",
    "correo_instancia",
]


def encontrar_excel_fuente() -> Path:
    """Busca el mismo archivo local que utiliza el dashboard."""
    candidatos = sorted(
        RAIZ_PROYECTO.glob("Fase_I_Evaluaci*n_360__180__90__copia_.xlsx"),
        key=lambda ruta: ruta.name.casefold(),
    )
    if not candidatos:
        raise FileNotFoundError(
            "No se encontró el Excel de Fase I en la raíz del proyecto. "
            "Indique la ruta con --entrada."
        )
    return candidatos[0]


def _primer_texto(fila: pd.Series, columnas: list[str]) -> str:
    for columna in columnas:
        valor = fila.get(columna)
        if pd.notna(valor) and str(valor).strip():
            return str(valor).strip()
    return ""


def _redondear_dos_decimales(valor: object) -> object:
    if pd.isna(valor):
        return pd.NA
    return float(Decimal(str(valor)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def construir_reporte(df_integrado: pd.DataFrame) -> pd.DataFrame:
    """Convierte la base integrada en una fila consolidada por persona."""
    if df_integrado.empty:
        return pd.DataFrame(columns=COLUMNAS_SALIDA)

    faltantes = [
        columna
        for columna in ["colaborador", "evd_360", "potencial", "objetivos"]
        if columna not in df_integrado.columns
    ]
    if faltantes:
        raise ValueError(
            "La base integrada no tiene las columnas requeridas: "
            + ", ".join(faltantes)
            + "."
        )

    reporte = pd.DataFrame(
        {
            "Nombre": df_integrado["colaborador"].fillna("").astype(str).str.strip(),
            "correo": df_integrado.apply(
                lambda fila: _primer_texto(fila, COLUMNAS_CORREO), axis=1
            ),
            "desempeño": pd.to_numeric(df_integrado["evd_360"], errors="coerce"),
            "competencias": pd.to_numeric(df_integrado["potencial"], errors="coerce"),
            "objetivos": pd.to_numeric(df_integrado["objetivos"], errors="coerce"),
        }
    )
    reporte = reporte[
        reporte["Nombre"].ne("") & reporte[COLUMNAS_PUNTAJE].notna().any(axis=1)
    ].copy()

    duplicados = reporte["Nombre"].str.casefold().duplicated(keep=False)
    if duplicados.any():
        muestra = ", ".join(
            reporte.loc[duplicados, "Nombre"].drop_duplicates().tolist()[:5]
        )
        raise ValueError(
            "La integración produjo nombres duplicados; revise las llaves de correo "
            f"o nombre para: {muestra}."
        )

    for columna in COLUMNAS_PUNTAJE:
        reporte[columna] = reporte[columna].map(_redondear_dos_decimales)
    return reporte.sort_values(
        "Nombre", key=lambda serie: serie.str.casefold()
    ).reset_index(drop=True)


def construir_resumen(ruta_entrada: Path) -> pd.DataFrame:
    """Carga las tres evaluaciones y construye el reporte general."""
    df_360 = motor_360.leer_exportacion_dashboard(ruta_entrada)
    metadata_organizacional = motor_360.leer_metadata_organizacional(ruta_entrada)
    if not metadata_organizacional.empty:
        df_360 = motor_360.completar_metadata_colaboradores(
            df_360,
            metadata_organizacional,
            "nombre_colaborador",
            "email_colaborador",
        )
    df_360_calculo = motor_360.filtrar_excluidos_desempeno(df_360)
    resultado_360 = motor_360.calcular_dashboard(
        df_360_calculo,
        motor_360.PESOS_BASE,
    )
    metadata_360 = motor_360.extraer_metadata_colaboradores(df_360)
    resultado_360["df_global"] = resultado_360["df_global"].merge(
        metadata_360,
        on="colaborador",
        how="left",
    )

    resultado_competencias = motor_competencias.leer_potencial(ruta_entrada)
    if not metadata_organizacional.empty:
        resultado_competencias["df_personas"] = motor_360.completar_metadata_colaboradores(
            resultado_competencias["df_personas"],
            metadata_organizacional,
            "colaborador",
            "correo",
        )
    resultado_objetivos = motor_objetivos.leer_objetivos(ruta_entrada)
    df_integrado = motor_integrado.preparar_resultado_integrado(
        resultado_360["df_global"],
        resultado_objetivos["df_colaboradores"],
        resultado_competencias["df_personas"],
        resultado_objetivos["df_fuente"],
        metadata_360,
    )
    return construir_reporte(df_integrado)


def guardar_excel(reporte: pd.DataFrame, ruta_salida: Path) -> None:
    """Escribe el reporte general como un libro Excel legible y filtrable."""
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Reporte general"
    hoja.sheet_view.showGridLines = False
    hoja.freeze_panes = "A2"

    hoja.append(COLUMNAS_SALIDA)
    for fila in reporte[COLUMNAS_SALIDA].itertuples(index=False, name=None):
        hoja.append([None if pd.isna(valor) else valor for valor in fila])

    relleno = PatternFill("solid", fgColor="185FA5")
    for celda in hoja[1]:
        celda.fill = relleno
        celda.font = Font(color="FFFFFF", bold=True)
        celda.alignment = Alignment(vertical="center")
    hoja.row_dimensions[1].height = 24

    hoja.column_dimensions["A"].width = 38
    hoja.column_dimensions["B"].width = 38
    for columna in ["C", "D", "E"]:
        hoja.column_dimensions[columna].width = 16
        for celda in hoja[columna][1:]:
            celda.number_format = "0.00"
            celda.alignment = Alignment(horizontal="right")

    ultima_fila = hoja.max_row
    if ultima_fila >= 2:
        tabla = Table(displayName="TablaReporteGeneral", ref=f"A1:E{ultima_fila}")
        tabla.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        hoja.add_table(tabla)

    libro.save(ruta_salida)


def parsear_argumentos() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Genera un Excel general con Nombre, correo, desempeño, "
            "competencias y objetivos."
        )
    )
    parser.add_argument(
        "--entrada",
        type=Path,
        help="Excel fuente. Si se omite, se busca el archivo usado por el dashboard.",
    )
    parser.add_argument(
        "--salida",
        type=Path,
        default=SALIDA_PREDETERMINADA,
        help=f"Ruta del Excel de salida (por defecto: {SALIDA_PREDETERMINADA}).",
    )
    return parser.parse_args()


def main() -> int:
    args = parsear_argumentos()
    entrada = (args.entrada or encontrar_excel_fuente()).resolve()
    salida = args.salida.resolve()

    if not entrada.is_file():
        raise FileNotFoundError(f"No existe el archivo de entrada: {entrada}")
    if entrada == salida:
        raise ValueError("La entrada y la salida no pueden ser el mismo archivo.")

    reporte = construir_resumen(entrada)
    guardar_excel(reporte, salida)
    print(f"Excel generado: {salida}")
    print(f"Personas exportadas: {len(reporte)}")
    for columna in COLUMNAS_PUNTAJE:
        print(f"Con {columna}: {int(reporte[columna].notna().sum())}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (FileNotFoundError, ValueError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1) from exc
